"""
Read Excel file content and output a human-readable table view.

Usage:
    python read_excel.py <file> [--sheet <name>] [--max-rows <n>]

Outputs a table with column-letter-prefixed headers, [空] marking empty cells,
and [合并] marking the non-top-left interior of merged ranges.
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def col_letter(col_idx: int) -> str:
    return get_column_letter(col_idx)


def cell_value_to_json(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return str(value)


def read_sheet(ws, max_rows: Optional[int] = None):
    merged = [str(r) for r in ws.merged_cells.ranges] if ws.merged_cells else []

    result = {
        "name": ws.title,
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
    }
    if merged:
        result["merged_cells"] = merged

    # Build set of merged-interior cells (non-top-left, cannot be written to)
    merged_interior = set()
    for mr in (ws.merged_cells.ranges if ws.merged_cells else []):
        first = True
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                if first:
                    first = False
                    continue
                merged_interior.add((row, col))

    # Detect header row
    header_row_idx = None
    headers = {}  # col_letter -> name
    for row_idx in range(1, min((ws.max_row or 0) + 1, 10)):
        non_empty = 0
        for col_idx in range(1, (ws.max_column or 0) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and str(val).strip():
                non_empty += 1
        if non_empty >= 2:
            header_row_idx = row_idx
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and str(val).strip():
                    headers[col_letter(col_idx)] = str(val).strip().replace('\n', ' ')
            break

    result["header_row"] = header_row_idx
    result["headers"] = headers

    # Read all rows
    end_row = ws.max_row or 0
    if max_rows and header_row_idx:
        end_row = min(end_row, header_row_idx + max_rows)

    rows = []
    for row_idx in range(1, end_row + 1):
        cells = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            col = col_letter(col_idx)
            if (row_idx, col_idx) in merged_interior:
                cells[col] = "__merged__"
                continue
            val = cell_value_to_json(ws.cell(row=row_idx, column=col_idx).value)
            cells[col] = val
        rows.append({"row": row_idx, "cells": cells})

    result["rows"] = rows
    return result


def format_table(sheet: dict) -> str:
    """Format sheet data as a human-readable table with [空] for empty cells."""
    lines = []
    headers = sheet.get("headers", {})
    header_row = sheet.get("header_row")
    rows = sheet.get("rows", [])
    merged = sheet.get("merged_cells", [])

    lines.append(f"Sheet: {sheet['name']} ({sheet.get('dimensions', '?')})")
    if merged:
        lines.append(f"Merged cells (do NOT write to non-top-left): {', '.join(merged)}")
    lines.append("")

    # Print pre-header rows (title, metadata)
    for row_data in rows:
        if header_row and row_data["row"] >= header_row:
            break
        vals = [v for v in row_data["cells"].values()
                if v is not None and v != "__merged__"]
        if vals:
            text = str(vals[0]).replace('\n', ' ')[:120]
            lines.append(text)
    lines.append("")

    if not headers:
        lines.append("(no header row detected)")
        return '\n'.join(lines)

    # Determine columns to show (only those with headers)
    cols = sorted(headers.keys())

    # Short header names with column letter prefix
    short_headers = {}
    for col, name in headers.items():
        short = name[:18] if len(name) > 18 else name
        short_headers[col] = f"{col}:{short}"

    # Build table header
    col_widths = {}
    for col in cols:
        col_widths[col] = max(len(short_headers[col]), 6)

    header_line = "row | " + " | ".join(
        f"{short_headers[col]:<{col_widths[col]}}" for col in cols
    )
    lines.append(header_line)
    lines.append("-" * len(header_line))

    # Data rows (after header)
    empty_count = 0
    for row_data in rows:
        if header_row and row_data["row"] <= header_row:
            continue
        cells = row_data["cells"]

        # Skip rows where all header columns are merged interior
        all_merged = all(cells.get(col) == "__merged__" for col in cols)
        if all_merged:
            continue

        parts = []
        row_has_empty = False
        for col in cols:
            val = cells.get(col)
            if val is None:
                parts.append(f"{'[空]':<{col_widths[col]}}")
                row_has_empty = True
                empty_count += 1
            elif val == "__merged__":
                parts.append(f"{'[合并]':<{col_widths[col]}}")
            else:
                s = str(val).replace('\n', ' ')
                if len(s) > col_widths[col]:
                    s = s[:col_widths[col] - 2] + ".."
                parts.append(f"{s:<{col_widths[col]}}")

        lines.append(f"{row_data['row']:>3} | " + " | ".join(parts))

    lines.append("")
    lines.append(f"Total empty cells (marked [空]): {empty_count}")
    return '\n'.join(lines)


def read_excel(file_path: str, sheet_name: Optional[str] = None, max_rows: Optional[int] = None):
    path = Path(file_path)
    if not path.exists():
        return {"error": f"File not found: {file_path}"}

    try:
        wb = load_workbook(str(path), data_only=True)
    except Exception as e:
        return {"error": f"Failed to open {file_path}: {e}"}

    result = {"file": str(path), "sheets": []}

    target_sheets = [sheet_name] if sheet_name else wb.sheetnames
    for name in target_sheets:
        if name not in wb.sheetnames:
            result["sheets"].append({"name": name, "error": f"Sheet '{name}' not found"})
            continue
        ws = wb[name]
        result["sheets"].append(read_sheet(ws, max_rows))

    wb.close()
    return result


def main():
    parser = argparse.ArgumentParser(description="Read Excel file content")
    parser.add_argument("file", help="Path to .xlsx/.xlsm file")
    parser.add_argument("--sheet", help="Read only this sheet (default: all)")
    parser.add_argument("--max-rows", type=int, default=None,
                        help="Max data rows to read per sheet (default: all)")
    args = parser.parse_args()

    result = read_excel(args.file, args.sheet, args.max_rows)

    if "error" in result:
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(1)

    for sheet in result.get("sheets", []):
        print(format_table(sheet))


if __name__ == "__main__":
    main()
