#!/usr/bin/env python3
import json
import sys

from openpyxl import load_workbook


def is_placeholder(value):
    if value is None:
        return True
    text = str(value).strip()
    if not text:
        return True
    return text.upper() in {"N/A", "NONE", "NULL", "UNNAMED"}


def is_numeric_like(value):
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    text = str(value).strip()
    if not text:
        return False
    if text.startswith("="):
        return True
    try:
        float(text)
        return True
    except ValueError:
        return False


def main():
    if len(sys.argv) != 2:
        print(
            json.dumps(
                {
                    "ok": False,
                    "error": "usage: verify_perf_review_fill.py <workbook_path>",
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 2

    path = sys.argv[1]
    wb = load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    failures = []

    if "绩效考核表" not in str(ws["A1"].value or ""):
        failures.append("title cell A1 was unexpectedly changed")
    if str(ws["A4"].value or "").strip() != "事项":
        failures.append("header cell A4 no longer contains '事项'")
    if "关键任务完成情况" not in str(ws["C5"].value or ""):
        failures.append("cell C5 no longer contains '关键任务完成情况'")
    if "代码质量" not in str(ws["C6"].value or ""):
        failures.append("cell C6 no longer contains '代码质量'")
    if str(ws["A12"].value or "").strip() != "总计":
        failures.append("summary cell A12 no longer contains '总计'")

    editable_filled = []
    for row in range(5, 12):
        for col in range(6, 12):
            cell = ws.cell(row=row, column=col)
            if not is_placeholder(cell.value):
                editable_filled.append(
                    {"coord": cell.coordinate, "value": str(cell.value).strip()}
                )

    numeric_scores = []
    for row in range(5, 12):
        for col in range(8, 11):
            cell = ws.cell(row=row, column=col)
            if is_numeric_like(cell.value):
                numeric_scores.append(
                    {"coord": cell.coordinate, "value": str(cell.value).strip()}
                )

    if len(editable_filled) < 12:
        failures.append(
            f"expected at least 12 filled editable cells, got {len(editable_filled)}"
        )
    if len(numeric_scores) < 6:
        failures.append(
            f"expected at least 6 numeric/formula score cells, got {len(numeric_scores)}"
        )

    report = {
        "ok": not failures,
        "path": path,
        "editable_filled_count": len(editable_filled),
        "numeric_score_count": len(numeric_scores),
        "editable_filled_samples": editable_filled[:20],
        "numeric_score_samples": numeric_scores[:20],
        "failures": failures,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
